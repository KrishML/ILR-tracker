import streamlit as st
import pandas as pd
from datetime import date, timedelta
import plotly.graph_objects as go
import openpyxl
import os

import db

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="UK ILR Absence Tracker", layout="wide")

# ── Constants ─────────────────────────────────────────────────────────────────
START_DATE    = "2022-10-22"
END_DATE      = "2027-11-24"
UKVI_LIMIT    = 180
EXCEL_FILE    = os.path.join(os.path.dirname(__file__), "ILR Vacation Tracker.xlsx")
EXCEL_SHEETS  = ["2022", "2023", "2024", "2025", "2026"]
TODAY         = date.today()

# ILR qualifying period anchor dates
# Per Appendix Continuous Residence CR 1.1 + CR 2.1(d):
# The qualifying period is based on continuous *residence* in the UK with permission.
# Time spent outside the UK before first entry cannot count as UK residence.
# Therefore the 5-year clock starts from the later of:
#   (a) visa grant date, or (b) first UK entry date
# Visa granted: 21 Oct 2022 | First UK entry: 24 Nov 2022
# → Qualifying period start = 24 Nov 2022 (entry date, as it is later)
# → Earliest possible ILR date = 24 Nov 2027
VISA_GRANT_DATE  = date(2022, 10, 21)   # for display only
UK_ENTRY_DATE    = date(2022, 11, 24)   # qualifying period anchor (later of grant/entry)
ILR_COMPLETION   = date(2027, 11, 24)   # UK_ENTRY_DATE + 5 years (qualifying period end)
# CR 1.1: applicant may apply up to 28 days before the 5-year completion date;
# the Home Office counts the qualifying period as ending on the most beneficial date
# (up to 28 days after application), effectively the completion date itself.
ILR_28DAY_EARLIEST = ILR_COMPLETION - timedelta(days=28)   # 27 Oct 2027
ILR_SEARCH_END   = ILR_COMPLETION   # end of tracked period


# ── DB init + Excel seed (runs once per process) ──────────────────────────────
@st.cache_resource
def initialise_db():
    """Create schema and seed Excel trips into the DB. Runs once per server process."""
    db.init_db()
    excel_trips = _read_excel_trips(EXCEL_FILE)
    seeded = db.seed_excel_trips(excel_trips)
    return seeded


@st.cache_data
def _read_excel_trips(path: str) -> list[tuple[str, str, str]]:
    """
    Read (destination, date_out, date_in) triples from each year sheet.
    Row layout: row 1 = title, row 2 = headers, rows 3+ = data.
    Columns: A=Country, B=Date Out, C=Date In, D=Duration (formula, ignored).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    trips: list[tuple[str, str, str]] = []
    for sheet_name in EXCEL_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in list(ws.iter_rows(values_only=True))[2:]:  # skip title + header
            dest, date_out, date_in = row[0], row[1], row[2]
            if isinstance(date_out, date) and isinstance(date_in, date):
                trips.append((
                    str(dest) if dest else "—",
                    date_out.strftime("%Y-%m-%d"),
                    date_in.strftime("%Y-%m-%d"),
                ))
    wb.close()
    return trips


initialise_db()


# ── Build absence dataframe ───────────────────────────────────────────────────
@st.cache_data
def build_dataframe(trips: tuple[tuple[str, str], ...]) -> pd.DataFrame:
    """
    Build a daily DataFrame with Absence and Rolling_365 columns.
    Accepts a tuple (hashable) so Streamlit can cache on it.
    UKVI whole-days rule: only days strictly between departure and return count.
    """
    df = pd.DataFrame({"Date": pd.date_range(start=START_DATE, end=END_DATE, freq="D")})
    df["Absence"] = 0

    for trip_start, trip_end in trips:
        ts = pd.Timestamp(trip_start)
        te = pd.Timestamp(trip_end)
        df.loc[(df["Date"] > ts) & (df["Date"] < te), "Absence"] = 1

    # Vectorised rolling 365-day sum
    df["Rolling_365"] = (
        df["Absence"]
        .rolling(window=365, min_periods=1)
        .sum()
        .astype(int)
    )
    return df


# ── Apply unplanned buffer days on top of a base dataframe ───────────────────
def apply_buffer_days(
    base_df: pd.DataFrame,
    extra_2026: int,
    extra_2027: int,
    start_2026: pd.Timestamp,
    start_2027: pd.Timestamp,
    end_2026: pd.Timestamp | None = None,
) -> pd.DataFrame:
    """
    Overlay slider buffer days onto a copy of base_df, then recompute Rolling_365.
    Days are added consecutively from the given start dates, skipping days already
    marked as absences. end_2026 caps the window for 2026 buffer days.
    """
    df = base_df.copy()

    def _fill(start: pd.Timestamp, n_days: int, end: pd.Timestamp | None = None) -> None:
        remaining = n_days
        for i in df.index:
            if remaining <= 0:
                break
            row_date = df.at[i, "Date"]
            if end is not None and row_date > end:
                break
            if row_date >= start and df.at[i, "Absence"] == 0:
                df.at[i, "Absence"] = 1
                remaining -= 1

    if extra_2026 > 0:
        _fill(start_2026, extra_2026, end_2026)
    if extra_2027 > 0:
        _fill(start_2027, extra_2027)

    df["Rolling_365"] = (
        df["Absence"].rolling(window=365, min_periods=1).sum().astype(int)
    )
    return df


# ── ILR eligibility assessment ────────────────────────────────────────────────
def _build_absence_series(
    date_range: pd.DatetimeIndex,
    all_trips_full: list[dict],
    extra_2026: int,
    extra_2027: int,
    buf_start_2026: pd.Timestamp,
    buf_end_2026: pd.Timestamp,
    buf_start_2027: pd.Timestamp,
) -> pd.Series:
    """
    Build a daily absence Series (0/1) for the given date_range,
    applying all DB trips plus slider buffer days.
    """
    absence = pd.Series(0, index=date_range, dtype=int)

    for t in all_trips_full:
        ts = pd.Timestamp(t["date_out"])
        te = pd.Timestamp(t["date_in"])
        mask = (date_range > ts) & (date_range < te)
        absence[mask] = 1

    def _fill(start: pd.Timestamp, end_cap: pd.Timestamp | None, n: int) -> None:
        remaining = n
        for d in date_range:
            if remaining <= 0:
                break
            if end_cap is not None and d > end_cap:
                break
            if d >= start and absence[d] == 0:
                absence[d] = 1
                remaining -= 1

    if extra_2026 > 0:
        _fill(buf_start_2026, buf_end_2026, extra_2026)
    if extra_2027 > 0:
        _fill(buf_start_2027, None, extra_2027)

    return absence


def assess_ilr(
    application_date: date,
    uk_entry_date: date,
    all_trips_full: list[dict],
    extra_2026: int,
    extra_2027: int,
    buf_start_2026: pd.Timestamp,
    buf_end_2026: pd.Timestamp,
    buf_start_2027: pd.Timestamp,
) -> dict:
    """
    Assess ILR eligibility under Appendix Continuous Residence (5-year Skilled Worker route).

    Rules checked:
      1. Qualifying period: 5 years of continuous UK residence ending on application_date.
         The period cannot start before uk_entry_date (first day of actual UK residence).
      2. No rolling 12-month window within that period may exceed 180 absence days
         (UKVI whole-days rule: departure and arrival days are NOT counted)
      3. No single trip may exceed 180 consecutive absence days
    """
    q_end   = pd.Timestamp(application_date)
    # CR 1.1: the Home Office uses whichever date is most beneficial — up to 28 days
    # after the application date. For a 28-day-early application this means the
    # qualifying period is assessed as if it ends on the 5-year completion date.
    completion_ts = pd.Timestamp(ILR_COMPLETION)
    if q_end < completion_ts and (completion_ts - q_end).days <= 28:
        q_end = completion_ts   # use completion date as the qualifying period end
    # Qualifying period starts 5 years before q_end, but no earlier than entry date
    q_start = max(q_end - pd.DateOffset(years=5), pd.Timestamp(uk_entry_date))
    date_range = pd.date_range(start=q_start, end=q_end, freq="D")

    absence = _build_absence_series(
        date_range, all_trips_full,
        extra_2026, extra_2027,
        buf_start_2026, buf_end_2026, buf_start_2027,
    )

    total_absences = int(absence.sum())

    breaches: list[dict] = []
    worst_days = 0
    worst_window: dict | None = None

    for i in range(len(date_range) - 365 + 1):
        days_absent = int(absence.iloc[i : i + 365].sum())
        if days_absent > worst_days:
            worst_days = days_absent
            worst_window = {
                "start": date_range[i].date(),
                "end":   date_range[i + 364].date(),
                "days":  days_absent,
            }
        if days_absent > UKVI_LIMIT:
            breaches.append({
                "start": date_range[i].date(),
                "end":   date_range[i + 364].date(),
                "days":  days_absent,
            })

    long_trips: list[dict] = []
    for t in all_trips_full:
        ts = pd.Timestamp(t["date_out"])
        te = pd.Timestamp(t["date_in"])
        if te < q_start or ts > q_end:
            continue
        consecutive = max((te - ts).days - 1, 0)
        if consecutive > UKVI_LIMIT:
            long_trips.append({
                "destination": t["destination"],
                "date_out": t["date_out"],
                "date_in":  t["date_in"],
                "days":     consecutive,
            })

    eligible = len(breaches) == 0 and len(long_trips) == 0

    return {
        "eligible":         eligible,
        "breaches":         breaches,
        "long_trips":       long_trips,
        "worst_window":     worst_window,
        "worst_days":       worst_days,
        "total_absences":   total_absences,
        "qualifying_start": q_start.date(),
        "qualifying_end":   q_end.date(),
    }


def compute_tight_windows(
    uk_entry_date: date,
    ilr_target_date: date,
    all_trips_full: list[dict],
    extra_2026: int,
    extra_2027: int,
    buf_start_2026: pd.Timestamp,
    buf_end_2026: pd.Timestamp,
    buf_start_2027: pd.Timestamp,
    warn_threshold: int = 100,
) -> list[dict]:
    """
    Find all calendar months in the future where the tightest rolling 12-month
    window that *includes* that month has fewer than (180 - warn_threshold) days
    of headroom (i.e. used > warn_threshold days).

    Returns a list of dicts sorted by headroom ascending:
      month_start, month_end, window_start, window_end,
      used, headroom, risk_level ('critical'|'high'|'moderate')
    """
    full_range = pd.date_range(
        start=pd.Timestamp(uk_entry_date),
        end=pd.Timestamp(ilr_target_date),
        freq="D",
    )
    absence = _build_absence_series(
        full_range, all_trips_full,
        extra_2026, extra_2027,
        buf_start_2026, buf_end_2026, buf_start_2027,
    )
    cum = absence.values.cumsum()

    def ws(s, e):
        return int(cum[e]) if s == 0 else int(cum[e] - cum[s - 1])

    date_to_idx = {d: i for i, d in enumerate(full_range)}
    today_ts    = pd.Timestamp(TODAY)

    # For each future calendar month, find the tightest 365-day window
    # that overlaps that month
    results = []
    check_start = pd.Timestamp(TODAY.replace(day=1))
    check_end   = pd.Timestamp(ilr_target_date)

    month_cursor = check_start
    while month_cursor <= check_end:
        m_start = month_cursor
        # last day of month
        if m_start.month == 12:
            m_end = pd.Timestamp(m_start.year + 1, 1, 1) - pd.Timedelta(days=1)
        else:
            m_end = pd.Timestamp(m_start.year, m_start.month + 1, 1) - pd.Timedelta(days=1)
        m_end = min(m_end, check_end)

        # Find the tightest (most-used) 365-day window that overlaps this month
        tightest_used  = 0
        tightest_ws    = None
        tightest_we    = None

        for i in range(len(full_range) - 364):
            w_start = full_range[i]
            w_end   = full_range[i + 364]
            # Window must overlap the month AND end in the future
            if w_end < today_ts:
                continue
            if w_end < m_start or w_start > m_end:
                continue
            used = ws(i, i + 364)
            if used > tightest_used:
                tightest_used  = used
                tightest_ws    = w_start.date()
                tightest_we    = w_end.date()

        if tightest_ws is not None:
            headroom = UKVI_LIMIT - tightest_used
            if headroom < (UKVI_LIMIT - warn_threshold):
                if headroom <= 50:
                    risk = "critical"
                elif headroom <= 70:
                    risk = "high"
                else:
                    risk = "moderate"
                results.append({
                    "month_label":  m_start.strftime("%B %Y"),
                    "month_start":  m_start.date(),
                    "month_end":    m_end.date(),
                    "window_start": tightest_ws,
                    "window_end":   tightest_we,
                    "used":         tightest_used,
                    "headroom":     headroom,
                    "risk":         risk,
                })

        # advance to next month
        if m_start.month == 12:
            month_cursor = pd.Timestamp(m_start.year + 1, 1, 1)
        else:
            month_cursor = pd.Timestamp(m_start.year, m_start.month + 1, 1)

    # Deduplicate by binding window (same window_start/end can appear for consecutive months)
    seen = set()
    deduped = []
    for r in sorted(results, key=lambda x: x["headroom"]):
        key = (r["window_start"], r["window_end"])
        if key not in seen:
            seen.add(key)
            deduped.append(r)

    return deduped


def find_earliest_ilr_date(
    uk_entry_date: date,
    all_trips_full: list[dict],
    extra_2026: int,
    extra_2027: int,
    buf_start_2026: pd.Timestamp,
    buf_end_2026: pd.Timestamp,
    buf_start_2027: pd.Timestamp,
    search_end: date,
) -> dict:
    """
    Scan forward day-by-day from (uk_entry_date + 5 years) to search_end,
    returning the first date where the 5-year qualifying window has no rolling
    12-month period exceeding 180 absence days.

    The qualifying period anchor is the UK entry date (first day of actual
    UK residence with permission), per Appendix Continuous Residence CR 1.1
    and CR 2.1(d). Time outside the UK before first entry does not count.

    Uses a cumulative sum for O(1) window queries.
    """
    # Earliest possible application = completion date − 28 days (CR 1.1 28-day rule)
    # The qualifying period is always assessed as ending on the completion date,
    # so we only need to check one effective q_end: ILR_COMPLETION.
    # We scan from 28 days before completion to search_end.
    earliest_possible = date(
        uk_entry_date.year + 5,
        uk_entry_date.month,
        uk_entry_date.day,
    ) - timedelta(days=28)

    if earliest_possible > search_end:
        return {"earliest_date": None, "days_to_go": None, "worst_days": None,
                "remaining_budget": None, "binding_w_start": None, "binding_w_end": None}

    # Build absence series from entry date to search_end
    full_start = pd.Timestamp(uk_entry_date)
    full_end   = pd.Timestamp(search_end)
    full_range = pd.date_range(start=full_start, end=full_end, freq="D")

    absence = _build_absence_series(
        full_range, all_trips_full,
        extra_2026, extra_2027,
        buf_start_2026, buf_end_2026, buf_start_2027,
    )

    # Pre-compute cumulative sum for O(1) window queries
    cum = absence.values.cumsum()

    def window_sum(start_idx: int, end_idx: int) -> int:
        if start_idx == 0:
            return int(cum[end_idx])
        return int(cum[end_idx] - cum[start_idx - 1])

    date_to_idx = {d: i for i, d in enumerate(full_range)}

    candidate = earliest_possible
    # The qualifying period always ends on ILR_COMPLETION (CR 1.1 most-beneficial-date rule).
    # So we only need to check one q_end. The scan finds the earliest *application* date
    # from which that qualifying window is breach-free.
    # Since the qualifying window is fixed, we check it once and return the earliest
    # application date (earliest_possible) if it passes, or None if it fails.
    q_end_ts    = pd.Timestamp(ILR_COMPLETION)
    q_start_ts  = pd.Timestamp(uk_entry_date)
    q_start_idx = date_to_idx.get(q_start_ts)
    q_end_idx   = date_to_idx.get(q_end_ts)

    while candidate <= search_end:
        # For each candidate application date, qualifying window is fixed:
        # q_start = uk_entry_date, q_end = ILR_COMPLETION (via 28-day rule)
        if q_start_idx is None or q_end_idx is None:
            candidate += timedelta(days=1)
            continue

        q_len = q_end_idx - q_start_idx + 1
        if q_len < 365:
            candidate += timedelta(days=1)
            continue

        # Check all 365-day windows within the qualifying period
        breach_found = False
        for i in range(q_start_idx, q_end_idx - 364 + 1):
            if window_sum(i, i + 364) > UKVI_LIMIT:
                breach_found = True
                break

        if not breach_found:
            # Check no single trip > 180 consecutive days
            long_trip = False
            for t in all_trips_full:
                ts = pd.Timestamp(t["date_out"])
                te = pd.Timestamp(t["date_in"])
                if te < q_start_ts or ts > q_end_ts:
                    continue
                if max((te - ts).days - 1, 0) > UKVI_LIMIT:
                    long_trip = True
                    break

            if not long_trip:
                days_to_go = (candidate - TODAY).days
                worst = max(
                    window_sum(i, i + 364)
                    for i in range(q_start_idx, q_end_idx - 364 + 1)
                )
                # Remaining travel budget
                today_idx = date_to_idx.get(pd.Timestamp(TODAY))
                remaining_budget = UKVI_LIMIT
                binding_w_start  = None
                binding_w_end    = None
                if today_idx is not None:
                    for j in range(len(full_range) - 364):
                        w_end_idx = j + 364
                        if w_end_idx < today_idx:
                            continue
                        h = UKVI_LIMIT - window_sum(j, w_end_idx)
                        if h < remaining_budget:
                            remaining_budget = h
                            binding_w_start  = full_range[j].date()
                            binding_w_end    = full_range[w_end_idx].date()

                return {
                    "earliest_date":    candidate,
                    "days_to_go":       days_to_go,
                    "worst_days":       worst,
                    "remaining_budget": remaining_budget,
                    "binding_w_start":  binding_w_start,
                    "binding_w_end":    binding_w_end,
                }

        # If breach found, no earlier application date will help (window is fixed)
        break

    return {"earliest_date": None, "days_to_go": None, "worst_days": None,
            "remaining_budget": None, "binding_w_start": None, "binding_w_end": None}

# ── Session state ─────────────────────────────────────────────────────────────
if "edit_id" not in st.session_state:
    st.session_state.edit_id = None   # DB id of the trip being edited


# ── Title ─────────────────────────────────────────────────────────────────────
st.title("🇬🇧 UK ILR Absence Tracker")
st.caption(
    "Tracks absences from the UK against the UKVI limit of **180 days** "
    "in any rolling 365-day period."
)

# ── Future trips manager ──────────────────────────────────────────────────────
st.subheader("Planned Future Travel")

# ── Add / Edit form ───────────────────────────────────────────────────────────
editing    = st.session_state.edit_id is not None
form_title = "✏️ Edit Trip" if editing else "➕ Add a Future Trip"

with st.expander(form_title, expanded=editing):
    if editing:
        # Load the trip being edited from the DB
        edit_trip = next(
            (t for t in db.get_manual_trips() if t["id"] == st.session_state.edit_id),
            None,
        )
        if edit_trip is None:
            st.session_state.edit_id = None
            st.rerun()
        default_dest     = edit_trip["destination"]
        default_date_out = date.fromisoformat(edit_trip["date_out"])
        default_date_in  = date.fromisoformat(edit_trip["date_in"])
    else:
        default_dest     = ""
        default_date_out = TODAY + timedelta(days=1)
        default_date_in  = TODAY + timedelta(days=8)

    with st.form("trip_form", clear_on_submit=True):
        fc1, fc2, fc3 = st.columns([2, 1, 1])
        with fc1:
            dest  = st.text_input("Destination", value=default_dest)
        with fc2:
            d_out = st.date_input("Date Out", value=default_date_out, key="form_out")
        with fc3:
            d_in  = st.date_input("Date In",  value=default_date_in,  key="form_in")

        fb1, fb2 = st.columns([1, 5])
        with fb1:
            submitted = st.form_submit_button("💾 Save" if editing else "➕ Add")
        with fb2:
            cancelled = st.form_submit_button("✖ Cancel") if editing else False

    if cancelled:
        st.session_state.edit_id = None
        st.rerun()

    if submitted:
        if d_out >= d_in:
            st.error("Date Out must be before Date In.")
        else:
            clean_dest = dest.strip() or "—"
            if editing:
                db.update_trip(st.session_state.edit_id, clean_dest, d_out.isoformat(), d_in.isoformat())
                st.session_state.edit_id = None
                st.success("Trip updated.")
            else:
                db.add_trip(clean_dest, d_out.isoformat(), d_in.isoformat())
                st.success(f"Trip to {clean_dest} added.")
            build_dataframe.clear()   # invalidate chart cache
            st.rerun()

# ── Manual trip list ──────────────────────────────────────────────────────────
manual_trips = db.get_manual_trips()

if manual_trips:
    header = st.columns([2, 1, 1, 1, 1, 1])
    for col, label in zip(header, ["Destination", "Date Out", "Date In", "Days Away", "", ""]):
        col.markdown(f"**{label}**")

    for trip in manual_trips:
        d_out_dt     = date.fromisoformat(trip["date_out"])
        d_in_dt      = date.fromisoformat(trip["date_in"])
        absence_days = max((d_in_dt - d_out_dt).days - 1, 0)

        row = st.columns([2, 1, 1, 1, 1, 1])
        row[0].write(trip["destination"])
        row[1].write(trip["date_out"])
        row[2].write(trip["date_in"])
        row[3].write(f"{absence_days}d")
        if row[4].button("✏️", key=f"edit_{trip['id']}", help="Edit"):
            st.session_state.edit_id = trip["id"]
            st.rerun()
        if row[5].button("🗑️", key=f"del_{trip['id']}", help="Delete"):
            db.delete_trip(trip["id"])
            build_dataframe.clear()
            st.rerun()
else:
    st.info("No future trips added yet. Use the form above to plan ahead.")

st.divider()

# ── Build base dataframe from all DB trips ────────────────────────────────────
all_trips = tuple(
    (t["date_out"], t["date_in"]) for t in db.get_all_trips()
)
df_base = build_dataframe(all_trips)

# ── Unplanned travel buffer sliders ──────────────────────────────────────────
st.subheader("Unplanned Travel Buffer")
st.caption(
    "Simulate extra unplanned days away to see how much flexibility remains "
    "before hitting the 180-day rolling limit."
)

# 2026 buffer: fixed window May 11 – Nov 12 2026 (gap between known trips)
# 2027 buffer: from day after the last known 2027 departure
BUFFER_2026_START = pd.Timestamp("2026-05-11")
BUFFER_2026_END   = pd.Timestamp("2026-11-20")   # inclusive upper bound for filling

def _last_departure_after(trips_tuple, year: int) -> pd.Timestamp:
    """Day after the latest departure date for trips departing in the given year."""
    outs = [
        pd.Timestamp(t[0])
        for t in trips_tuple
        if pd.Timestamp(t[0]).year == year
    ]
    return (max(outs) + pd.Timedelta(days=1)) if outs else pd.Timestamp(f"{year}-01-01")

start_2027 = _last_departure_after(all_trips, 2027)

col_s1, col_s2 = st.columns(2)
with col_s1:
    extra_2026 = st.slider(
        "Extra 2026 days (11 May – 20 Nov 2026)",
        min_value=0, max_value=120, value=0,
        key="slider_2026",
    )
with col_s2:
    extra_2027 = st.slider(
        f"Extra 2027 days (from {start_2027.strftime('%d %b %Y')})",
        min_value=0, max_value=180, value=0,
        key="slider_2027",
    )

# Apply buffer on top of the cached base (not persisted, purely for display)
df = apply_buffer_days(df_base, extra_2026, extra_2027, BUFFER_2026_START, start_2027, BUFFER_2026_END)

st.divider()

# ── Derived metrics ───────────────────────────────────────────────────────────
max_peak = int(df["Rolling_365"].max())

if max_peak < 150:
    status_label       = "✅ Safe"
    status_delta_color = "normal"
elif max_peak <= UKVI_LIMIT:
    status_label       = "⚠️ Warning Zone"
    status_delta_color = "off"
else:
    status_label       = "🚨 LIMIT EXCEEDED"
    status_delta_color = "inverse"

# ── Summary metrics ───────────────────────────────────────────────────────────
st.subheader("Summary")
m1, m2, m3 = st.columns(3)
m1.metric("UKVI Limit", f"{UKVI_LIMIT} Days")
m2.metric(
    "Max Rolling Peak",
    f"{max_peak} Days",
    delta=f"{max_peak - UKVI_LIMIT:+d} vs limit",
    delta_color=status_delta_color,
)
m3.metric("Status", status_label)

st.divider()

# ── Plotly chart ──────────────────────────────────────────────────────────────
fig = go.Figure()

fig.add_trace(go.Scatter(
    x=df["Date"],
    y=df["Rolling_365"],
    mode="lines",
    name="Rolling 365-day absences",
    line=dict(color="royalblue", width=2),
    fill="tozeroy",
    fillcolor="rgba(65, 105, 225, 0.15)",
))

fig.add_hline(
    y=UKVI_LIMIT,
    line_dash="dash",
    line_color="red",
    line_width=2,
    annotation_text="UKVI 180-day limit",
    annotation_position="top left",
    annotation_font_color="red",
)

fig.update_layout(
    title="Rolling 365-Day UK Absences",
    xaxis_title="Date",
    yaxis_title="Days Absent (rolling 365-day window)",
    yaxis=dict(range=[0, 200]),
    hovermode="x unified",
    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    margin=dict(l=40, r=40, t=60, b=80),
    template="plotly_white",
    height=600,
    xaxis=dict(
        rangeslider=dict(visible=True, thickness=0.05),
        type="date",
        rangeselector=dict(buttons=[
            dict(count=1,  label="1m",  step="month"),
            dict(count=3,  label="3m",  step="month"),
            dict(count=6,  label="6m",  step="month"),
            dict(count=1,  label="YTD", step="year", stepmode="todate"),
            dict(step="all", label="All"),
        ]),
    ),
)

st.plotly_chart(fig, use_container_width=True)

# ── Trip history by year ──────────────────────────────────────────────────────
st.subheader("Trip History")

all_trips_full = db.get_all_trips()   # full dicts with destination + source

# Group by departure year
from collections import defaultdict
trips_by_year: dict[int, list[dict]] = defaultdict(list)
for t in all_trips_full:
    year = int(t["date_out"][:4])
    trips_by_year[year].append(t)

YEAR_COLORS = {
    2022: "#6366f1",   # indigo
    2023: "#0ea5e9",   # sky blue
    2024: "#10b981",   # emerald
    2025: "#f59e0b",   # amber
    2026: "#ef4444",   # red
    2027: "#8b5cf6",   # violet
}

for year in sorted(trips_by_year.keys(), reverse=True):
    year_trips = sorted(trips_by_year[year], key=lambda t: t["date_out"])
    total_absence = sum(
        max((date.fromisoformat(t["date_in"]) - date.fromisoformat(t["date_out"])).days - 1, 0)
        for t in year_trips
    )
    color = YEAR_COLORS.get(year, "#64748b")

    # Year header with total
    st.markdown(
        f"""
        <div style="display:flex; align-items:center; gap:12px; margin: 20px 0 8px 0;">
            <div style="background:{color}; color:white; font-weight:700; font-size:1.1rem;
                        padding:4px 16px; border-radius:20px;">{year}</div>
            <div style="color:#64748b; font-size:0.9rem;">
                {len(year_trips)} trip{"s" if len(year_trips) != 1 else ""} &nbsp;·&nbsp;
                <strong>{total_absence}</strong> absence days
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Trip cards — 3 per row
    cols_per_row = 3
    for row_start in range(0, len(year_trips), cols_per_row):
        row_trips = year_trips[row_start : row_start + cols_per_row]
        cols = st.columns(cols_per_row)
        for col, trip in zip(cols, row_trips):
            d_out = date.fromisoformat(trip["date_out"])
            d_in  = date.fromisoformat(trip["date_in"])
            absence_days = max((d_in - d_out).days - 1, 0)
            total_days   = (d_in - d_out).days

            source_badge = (
                f'<span style="background:#e0f2fe; color:#0369a1; font-size:0.7rem; '
                f'padding:2px 8px; border-radius:10px; font-weight:600;">PLANNED</span>'
                if trip["source"] == "manual"
                else f'<span style="background:#f0fdf4; color:#166534; font-size:0.7rem; '
                f'padding:2px 8px; border-radius:10px; font-weight:600;">EXCEL</span>'
            )

            col.markdown(
                f"""
                <div style="border:1px solid #e2e8f0; border-radius:12px; padding:14px 16px;
                            border-left:4px solid {color}; background:#fafafa;
                            margin-bottom:8px; min-height:110px;">
                    <div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:6px;">
                        <div style="font-weight:600; font-size:0.95rem; color:#1e293b;
                                    flex:1; margin-right:8px;">{trip["destination"]}</div>
                        {source_badge}
                    </div>
                    <div style="color:#475569; font-size:0.82rem; margin-bottom:8px;">
                        📅 {d_out.strftime("%d %b")} → {d_in.strftime("%d %b %Y")}
                    </div>
                    <div style="display:flex; gap:8px;">
                        <span style="background:{color}22; color:{color}; font-size:0.78rem;
                                     padding:2px 10px; border-radius:10px; font-weight:600;">
                            {absence_days}d absent
                        </span>
                        <span style="background:#f1f5f9; color:#64748b; font-size:0.78rem;
                                     padding:2px 10px; border-radius:10px;">
                            {total_days}d total
                        </span>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    st.markdown("<div style='margin-bottom:4px'></div>", unsafe_allow_html=True)

st.divider()

# ── Earliest ILR Date Prediction ──────────────────────────────────────────────
st.subheader("📅 Earliest ILR Application Date")
st.caption(
    "Under CR 1.1 (Appendix Continuous Residence), you may apply up to **28 days before** "
    "your 5-year completion date. The Home Office assesses the qualifying period as ending "
    "on the completion date (most beneficial date rule). "
    "Your 5-year completion date is **24 Nov 2027** (5 years from first UK entry). "
    "Earliest application: **27 Oct 2027**. Updates live with every trip change."
)

with st.spinner("Calculating earliest eligible date…"):
    prediction = find_earliest_ilr_date(
        uk_entry_date=UK_ENTRY_DATE,
        all_trips_full=all_trips_full,
        extra_2026=extra_2026,
        extra_2027=extra_2027,
        buf_start_2026=BUFFER_2026_START,
        buf_end_2026=BUFFER_2026_END,
        buf_start_2027=start_2027,
        search_end=ILR_SEARCH_END,
    )

if prediction["earliest_date"] is None:
    st.markdown(
        """
        <div style="background:#fef2f2; border:1.5px solid #fca5a5; border-radius:12px;
                    padding:20px 24px;">
            <div style="font-size:1.2rem; font-weight:700; color:#dc2626; margin-bottom:4px;">
                ⚠️ No eligible date found within the tracked period
            </div>
            <div style="color:#991b1b; font-size:0.9rem;">
                Based on current and projected absences, no date before
                24 Nov 2027 satisfies the continuous residence requirement.
                Consider reducing planned travel.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
else:
    ed          = prediction["earliest_date"]
    days_to_go  = prediction["days_to_go"]
    worst       = prediction["worst_days"]
    headroom    = UKVI_LIMIT - worst

    # Countdown label
    if days_to_go < 0:
        countdown_html = (
            f'<span style="color:#15803d; font-weight:600;">Already passed '
            f'({abs(days_to_go)} days ago)</span>'
        )
    elif days_to_go == 0:
        countdown_html = '<span style="color:#15803d; font-weight:600;">Today!</span>'
    else:
        years_left  = days_to_go // 365
        months_left = (days_to_go % 365) // 30
        days_left   = days_to_go % 30
        parts = []
        if years_left:  parts.append(f"{years_left}y")
        if months_left: parts.append(f"{months_left}m")
        if days_left:   parts.append(f"{days_left}d")
        countdown_html = (
            f'<span style="color:#0369a1; font-weight:600;">'
            f'{"  ".join(parts)} from today</span>'
        )

    headroom_color = "#15803d" if headroom >= 30 else "#d97706" if headroom >= 10 else "#dc2626"

    st.markdown(
        f"""
        <div style="background:linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
                    border:1.5px solid #7dd3fc; border-radius:14px;
                    padding:24px 28px; margin-bottom:8px;">
            <div style="font-size:0.85rem; color:#0369a1; font-weight:600;
                        text-transform:uppercase; letter-spacing:0.05em; margin-bottom:6px;">
                Earliest Eligible ILR Application Date
            </div>
            <div style="font-size:2.2rem; font-weight:800; color:#0c4a6e; margin-bottom:4px;">
                {ed.strftime("%d %B %Y")}
            </div>
            <div style="font-size:1rem; margin-bottom:16px;">
                {countdown_html}
            </div>
            <div style="display:flex; gap:24px; flex-wrap:wrap;">
                <div>
                    <div style="font-size:0.75rem; color:#64748b; text-transform:uppercase;">
                        Qualifying Period
                    </div>
                    <div style="font-weight:600; color:#1e293b; font-size:0.9rem;">
                        {UK_ENTRY_DATE.strftime("%d %b %Y")} → {ILR_COMPLETION.strftime("%d %b %Y")}
                    </div>
                </div>
                <div>
                    <div style="font-size:0.75rem; color:#64748b; text-transform:uppercase;">
                        5-Year Completion Date
                    </div>
                    <div style="font-weight:600; color:#1e293b; font-size:0.9rem;">
                        {ILR_COMPLETION.strftime("%d %b %Y")}
                    </div>
                </div>
                <div>
                    <div style="font-size:0.75rem; color:#64748b; text-transform:uppercase;">
                        Worst Rolling Window
                    </div>
                    <div style="font-weight:600; color:#1e293b; font-size:0.9rem;">
                        {worst} / 180 days
                    </div>
                </div>
                <div>
                    <div style="font-size:0.75rem; color:#64748b; text-transform:uppercase;">
                        Headroom
                    </div>
                    <div style="font-weight:700; color:{headroom_color}; font-size:0.9rem;">
                        {headroom} days to spare
                    </div>
                </div>
                <div>
                    <div style="font-size:0.75rem; color:#64748b; text-transform:uppercase;">
                        Visa Granted / UK Entry
                    </div>
                    <div style="font-weight:600; color:#1e293b; font-size:0.9rem;">
                        {VISA_GRANT_DATE.strftime("%d %b %Y")} / {UK_ENTRY_DATE.strftime("%d %b %Y")}
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

st.divider()

# ── Travel Caution Periods ────────────────────────────────────────────────────
st.subheader("🚦 Travel Caution Periods")
st.caption(
    "Future calendar months where a rolling 12-month window is already tight. "
    "Adding travel in these periods risks pushing a window over the 180-day limit "
    "and could delay your ILR eligibility date. Updates live with every trip change."
)

tight_windows = compute_tight_windows(
    uk_entry_date=UK_ENTRY_DATE,
    ilr_target_date=ILR_SEARCH_END,
    all_trips_full=all_trips_full,
    extra_2026=extra_2026,
    extra_2027=extra_2027,
    buf_start_2026=BUFFER_2026_START,
    buf_end_2026=BUFFER_2026_END,
    buf_start_2027=start_2027,
    warn_threshold=100,   # flag windows with < 80 days headroom (used > 100)
)

if not tight_windows:
    st.success("No tight windows detected — all future rolling periods have comfortable headroom.")
else:
    RISK_CONFIG = {
        "critical": {"color": "#dc2626", "bg": "#fef2f2", "border": "#fca5a5",
                     "icon": "🔴", "label": "CRITICAL"},
        "high":     {"color": "#d97706", "bg": "#fffbeb", "border": "#fcd34d",
                     "icon": "🟠", "label": "HIGH RISK"},
        "moderate": {"color": "#0369a1", "bg": "#f0f9ff", "border": "#7dd3fc",
                     "icon": "🟡", "label": "CAUTION"},
    }

    # Summary counts
    n_crit = sum(1 for w in tight_windows if w["risk"] == "critical")
    n_high = sum(1 for w in tight_windows if w["risk"] == "high")
    n_mod  = sum(1 for w in tight_windows if w["risk"] == "moderate")

    summary_parts = []
    if n_crit: summary_parts.append(f"🔴 **{n_crit} critical**")
    if n_high: summary_parts.append(f"🟠 **{n_high} high risk**")
    if n_mod:  summary_parts.append(f"🟡 **{n_mod} caution**")
    st.markdown("  ·  ".join(summary_parts) + "  windows detected")

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    cols_per_row = 3
    for row_start in range(0, len(tight_windows), cols_per_row):
        row_items = tight_windows[row_start : row_start + cols_per_row]
        cols = st.columns(cols_per_row)
        for col, w in zip(cols, row_items):
            cfg = RISK_CONFIG[w["risk"]]
            bar_pct = min(int(w["used"] / UKVI_LIMIT * 100), 100)
            bar_color = cfg["color"]
            col.markdown(
                f"""
                <div style="background:{cfg['bg']}; border:1.5px solid {cfg['border']};
                            border-radius:12px; padding:14px 16px; margin-bottom:10px;">
                    <div style="display:flex; justify-content:space-between; align-items:center;
                                margin-bottom:8px;">
                        <div style="font-weight:700; font-size:0.95rem; color:{cfg['color']};">
                            {cfg['icon']} {w['month_label']}
                        </div>
                        <span style="background:{cfg['color']}22; color:{cfg['color']};
                                     font-size:0.7rem; padding:2px 8px; border-radius:8px;
                                     font-weight:700;">{cfg['label']}</span>
                    </div>
                    <div style="font-size:0.78rem; color:#475569; margin-bottom:10px;">
                        Binding window:<br>
                        <strong>{w['window_start'].strftime('%d %b %Y')}
                        → {w['window_end'].strftime('%d %b %Y')}</strong>
                    </div>
                    <div style="background:#e2e8f0; border-radius:6px;
                                height:8px; margin-bottom:6px; overflow:hidden;">
                        <div style="background:{bar_color}; width:{bar_pct}%;
                                    height:100%; border-radius:6px;"></div>
                    </div>
                    <div style="display:flex; justify-content:space-between;
                                font-size:0.78rem;">
                        <span style="color:#64748b;">{w['used']} / 180 days used</span>
                        <span style="font-weight:700; color:{cfg['color']};">
                            {w['headroom']}d left
                        </span>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

st.divider()

# ── ILR Eligibility Assessment ────────────────────────────────────────────────
st.subheader("🏛️ ILR Eligibility Assessment")
st.caption(
    "Assesses eligibility under the **5-year Skilled Worker route** "
    "(Appendix Continuous Residence). The qualifying period runs from your "
    "first UK entry date (24 Nov 2022) to your chosen application date. "
    "The 180-day rule is checked across every rolling 12-month window in that period."
)

ILR_MIN_DATE = ILR_28DAY_EARLIEST   # 27 Oct 2027 — earliest application under 28-day rule
ILR_MAX_DATE = ILR_COMPLETION       # 24 Nov 2027 — 5-year completion date

ilr_date = st.date_input(
    "Select ILR application date",
    value=ILR_28DAY_EARLIEST,
    min_value=ILR_MIN_DATE,
    max_value=ILR_MAX_DATE,
    format="DD/MM/YYYY",
    key="ilr_date",
)

if ilr_date < ILR_MIN_DATE:
    st.error(f"ILR application date must be on or after {ILR_MIN_DATE.strftime('%d %b %Y')}.")
else:
    result = assess_ilr(
        application_date=ilr_date,
        uk_entry_date=UK_ENTRY_DATE,
        all_trips_full=all_trips_full,
        extra_2026=extra_2026,
        extra_2027=extra_2027,
        buf_start_2026=BUFFER_2026_START,
        buf_end_2026=BUFFER_2026_END,
        buf_start_2027=start_2027,
    )

    # ── Verdict banner ────────────────────────────────────────────────────────
    if result["eligible"]:
        st.markdown(
            """
            <div style="background:#f0fdf4; border:1.5px solid #86efac; border-radius:12px;
                        padding:20px 24px; margin-bottom:16px;">
                <div style="font-size:1.5rem; font-weight:700; color:#15803d; margin-bottom:4px;">
                    ✅ Eligible for ILR
                </div>
                <div style="color:#166534; font-size:0.95rem;">
                    No rolling 12-month window exceeds 180 absence days during the qualifying period.
                    You appear to meet the continuous residence requirement.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            """
            <div style="background:#fef2f2; border:1.5px solid #fca5a5; border-radius:12px;
                        padding:20px 24px; margin-bottom:16px;">
                <div style="font-size:1.5rem; font-weight:700; color:#dc2626; margin-bottom:4px;">
                    ❌ Not Eligible — Continuous Residence Broken
                </div>
                <div style="color:#991b1b; font-size:0.95rem;">
                    One or more rolling 12-month windows exceed the 180-day absence limit.
                    Continuous residence is considered broken under Appendix Continuous Residence.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ── Key figures ───────────────────────────────────────────────────────────
    a1, a2, a3, a4 = st.columns(4)
    a1.metric("Qualifying Period Start", result["qualifying_start"].strftime("%d %b %Y"))
    a2.metric("Qualifying Period End",   result["qualifying_end"].strftime("%d %b %Y"))
    a3.metric("Total Absence Days",      f"{result['total_absences']}d")
    a4.metric(
        "Worst Rolling Window",
        f"{result['worst_days']}d",
        delta=f"{result['worst_days'] - UKVI_LIMIT:+d} vs 180-day limit",
        delta_color="inverse" if result["worst_days"] > UKVI_LIMIT else "normal",
    )

    # ── Worst window detail ───────────────────────────────────────────────────
    if result["worst_window"]:
        ww = result["worst_window"]
        status_color = "#dc2626" if ww["days"] > UKVI_LIMIT else "#15803d"
        st.markdown(
            f"""
            <div style="background:#f8fafc; border:1px solid #e2e8f0; border-radius:10px;
                        padding:14px 18px; margin-top:12px;">
                <span style="font-weight:600; color:#334155;">Peak absence window: </span>
                <span style="color:#475569;">
                    {ww["start"].strftime("%d %b %Y")} → {ww["end"].strftime("%d %b %Y")}
                </span>
                &nbsp;·&nbsp;
                <span style="font-weight:700; color:{status_color};">{ww["days"]} days absent</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ── Breach details ────────────────────────────────────────────────────────
    if result["breaches"]:
        with st.expander(f"⚠️ {len(result['breaches'])} breach window(s) — click to expand"):
            st.markdown(
                "<div style='color:#64748b; font-size:0.85rem; margin-bottom:8px;'>"
                "Each row is a 365-day window where absences exceeded 180 days.</div>",
                unsafe_allow_html=True,
            )
            breach_df = pd.DataFrame(result["breaches"])
            breach_df.columns = ["Window Start", "Window End", "Days Absent"]
            breach_df["Over Limit By"] = breach_df["Days Absent"] - UKVI_LIMIT
            st.dataframe(breach_df, use_container_width=True, hide_index=True)

    # ── Long single-trip warnings ─────────────────────────────────────────────
    if result["long_trips"]:
        with st.expander(f"⚠️ {len(result['long_trips'])} trip(s) exceeding 180 consecutive days"):
            for lt in result["long_trips"]:
                st.markdown(
                    f"- **{lt['destination']}** &nbsp; {lt['date_out']} → {lt['date_in']} "
                    f"&nbsp; ({lt['days']} consecutive absence days)",
                    unsafe_allow_html=True,
                )

    # ── Disclaimer ────────────────────────────────────────────────────────────
    st.markdown(
        "<div style='color:#94a3b8; font-size:0.78rem; margin-top:16px;'>"
        "⚠️ This assessment is for informational purposes only and does not constitute legal advice. "
        "Always consult a qualified immigration solicitor before submitting an ILR application."
        "</div>",
        unsafe_allow_html=True,
    )
